"""
Export utilities for generating reports and exports.
"""
import io
import csv
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import get_template
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class ScheduleExporter:
    """
    Export schedules to various formats.
    """
    
    @staticmethod
    def export_to_pdf(schedules, title="Emploi du temps"):
        """
        Export schedules to PDF format.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 20))
        
        # Table data
        data = [['Matière', 'Enseignant', 'Salle', 'Créneau', 'Date', 'Filières']]
        
        for schedule in schedules:
            programs = ', '.join([sp.program.name for sp in schedule.scheduleprogram_set.all()])
            data.append([
                schedule.subject.name,
                schedule.teacher.user.get_full_name(),
                schedule.room.name,
                str(schedule.time_slot),
                f"{schedule.start_date} - {schedule.end_date}",
                programs
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_to_excel(schedules, title="Emploi du temps"):
        """
        Export schedules to Excel format.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Emploi du temps"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ['Matière', 'Enseignant', 'Salle', 'Créneau', 'Date début', 'Date fin', 'Filières', 'Statut']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data
        for row, schedule in enumerate(schedules, 2):
            programs = ', '.join([sp.program.name for sp in schedule.scheduleprogram_set.all()])
            status = "Annulé" if schedule.is_cancelled else "Actif"
            
            ws.cell(row=row, column=1, value=schedule.subject.name)
            ws.cell(row=row, column=2, value=schedule.teacher.user.get_full_name())
            ws.cell(row=row, column=3, value=schedule.room.name)
            ws.cell(row=row, column=4, value=str(schedule.time_slot))
            ws.cell(row=row, column=5, value=schedule.start_date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=6, value=schedule.end_date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=7, value=programs)
            ws.cell(row=row, column=8, value=status)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_to_csv(schedules):
        """
        Export schedules to CSV format.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow(['Matière', 'Enseignant', 'Salle', 'Créneau', 'Date début', 'Date fin', 'Filières', 'Statut'])
        
        # Data
        for schedule in schedules:
            programs = ', '.join([sp.program.name for sp in schedule.scheduleprogram_set.all()])
            status = "Annulé" if schedule.is_cancelled else "Actif"
            
            writer.writerow([
                schedule.subject.name,
                schedule.teacher.user.get_full_name(),
                schedule.room.name,
                str(schedule.time_slot),
                schedule.start_date.strftime('%d/%m/%Y'),
                schedule.end_date.strftime('%d/%m/%Y'),
                programs,
                status
            ])
        
        output.seek(0)
        return output


class TeacherWorkloadExporter:
    """
    Export teacher workload reports.
    """
    
    @staticmethod
    def export_to_pdf(teachers):
        """
        Export teacher workload to PDF.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Charge de travail des enseignants", title_style))
        story.append(Spacer(1, 20))
        
        # Table data
        data = [['Enseignant', 'Département', 'Heures/semaine', 'Nombre de cours', 'Taux d\'occupation']]
        
        for teacher in teachers:
            weekly_hours = 20  # Mock data
            course_count = 5   # Mock data
            occupation_rate = f"{(weekly_hours / teacher.max_hours_per_week * 100):.1f}%"
            
            data.append([
                teacher.user.get_full_name(),
                teacher.department.name if hasattr(teacher, 'department') else 'N/A',
                f"{weekly_hours}h",
                str(course_count),
                occupation_rate
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_to_excel(teachers):
        """
        Export teacher workload to Excel.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Charge de travail"
        
        # Headers
        headers = ['Enseignant', 'Département', 'Heures/semaine', 'Nombre de cours', 'Taux d\'occupation']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row, teacher in enumerate(teachers, 2):
            weekly_hours = 20  # Mock data
            course_count = 5   # Mock data
            occupation_rate = weekly_hours / teacher.max_hours_per_week * 100
            
            ws.cell(row=row, column=1, value=teacher.user.get_full_name())
            ws.cell(row=row, column=2, value=getattr(teacher, 'department', {}).get('name', 'N/A'))
            ws.cell(row=row, column=3, value=weekly_hours)
            ws.cell(row=row, column=4, value=course_count)
            ws.cell(row=row, column=5, value=f"{occupation_rate:.1f}%")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_to_csv(teachers):
        """
        Export teacher workload to CSV.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['Enseignant', 'Département', 'Heures/semaine', 'Nombre de cours', 'Taux d\'occupation'])
        
        for teacher in teachers:
            weekly_hours = 20  # Mock data
            course_count = 5   # Mock data
            occupation_rate = f"{(weekly_hours / teacher.max_hours_per_week * 100):.1f}%"
            
            writer.writerow([
                teacher.user.get_full_name(),
                getattr(teacher, 'department', {}).get('name', 'N/A'),
                f"{weekly_hours}h",
                str(course_count),
                occupation_rate
            ])
        
        output.seek(0)
        return output


def generate_schedule_pdf_response(schedules, filename="emploi_du_temps.pdf"):
    """
    Generate HTTP response with PDF schedule.
    """
    buffer = ScheduleExporter.export_to_pdf(schedules)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_schedule_excel_response(schedules, filename="emploi_du_temps.xlsx"):
    """
    Generate HTTP response with Excel schedule.
    """
    buffer = ScheduleExporter.export_to_excel(schedules)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_schedule_csv_response(schedules, filename="emploi_du_temps.csv"):
    """
    Generate HTTP response with CSV schedule.
    """
    buffer = ScheduleExporter.export_to_csv(schedules)
    
    response = HttpResponse(buffer.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response